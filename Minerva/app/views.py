from rest_framework import viewsets
from django.contrib.auth.models import User
from .models import (
    UserProfile, DiarioDigital, Categoria,
    LinkRelevante, Articulo
)
from .serializers import (
    UserSerializer, UserProfileSerializer,
    DiarioDigitalSerializer, CategoriaSerializer,
    LinkRelevanteSerializer, ArticuloSerializer, LinkRelevanteCreateSerializer
)

from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from .models import Roles

from django.contrib.auth import authenticate, login
from django.shortcuts import render, redirect
from django.contrib.auth.forms import AuthenticationForm

from django.contrib.auth import logout
from django.shortcuts import redirect

from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator
from .models import Roles

from rest_framework import viewsets
from .models import LinkRelevante
from .serializers import LinkRelevanteSerializer
from django.utils.dateparse import parse_date

from .utils.actividad import log_actividad
from .models import TipoActividad

from rest_framework import viewsets
from django.utils.dateparse import parse_date
from .models import Actividad
from .serializers import ActividadSerializer

import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from .models import Actividad
from django.utils.dateparse import parse_date

from rest_framework import viewsets
from rest_framework.permissions import IsAuthenticated
from django.utils.dateparse import parse_date

from .models import LinkRelevante
from .serializers import (
    LinkRelevanteSerializer,
    LinkRelevanteCreateSerializer
)

from django.views.decorators.csrf import csrf_exempt
import json
from django.http import JsonResponse

def login_view(request):
    if request.method == 'POST':
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            user = form.get_user()
            login(request, user)
            log_actividad(request, TipoActividad.LOGIN, "Inicio de sesión exitoso")
            rol = getattr(user.userprofile, 'rol', None)
            if rol == Roles.PRENSA:
                return redirect('prensa')
            elif rol == Roles.CLASIFICACION:
                return redirect('clasificacion')
            elif rol == Roles.REDACCION:
                return redirect('redaccion')
            elif rol == Roles.ADMIN:
                return redirect('actividad')
            else:
                return render(request, '403.html', status=403)
    else:
        form = AuthenticationForm()
    return render(request, 'login.html', {'form': form})


def logout_view(request):
    log_actividad(request, TipoActividad.LOGOUT, "Cierre de sesión exitoso")
    logout(request)    
    return redirect('login')


@login_required
def prensa_view(request):
    user_profile = getattr(request.user, 'userprofile', None)
    if user_profile and user_profile.rol in [Roles.PRENSA, Roles.ADMIN]:
        return render(request, 'prensa.html')
    return render(request, '403.html', status=403)  # O redirigí a login o inicio

@login_required
def actividad_debug_view(request):
    return render(request, 'actividad_debug.html')


#----------------------------------------------------------------------------

@login_required
def clasificacion_view(request):
    user_profile = getattr(request.user, 'userprofile', None)
    if user_profile and user_profile.rol in [Roles.CLASIFICACION, Roles.ADMIN]:
        return render(request, 'clasificacion.html')
    return render(request, '403.html', status=403)  # O redirigí a login o inicio

#-----------------------------------------------------------------------------

@login_required
def redaccion_view(request):
    user_profile = getattr(request.user, 'userprofile', None)
    if user_profile and user_profile.rol in [Roles.PRENSA, Roles.ADMIN]:
        return render(request, 'redaccion.html')
    return render(request, '403.html', status=403)


#-------------------------------------------------------------------------------

from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from .models import Actividad, Roles

@login_required
def actividad_view(request):
    user_profile = getattr(request.user, 'userprofile', None)
    if user_profile and user_profile.rol == Roles.ADMIN:
        return render(request, 'actividad.html')
    return render(request, '403.html', status=403)

#---------------------------------------------------------------------------------


from django.http import HttpResponse
import openpyxl
from openpyxl.utils import get_column_letter
from .models import Actividad
from django.utils.dateparse import parse_date

def exportar_actividades_excel(request):
    queryset = Actividad.objects.select_related('usuario').order_by('-fecha_hora')

    # Filtros
    usuario = request.GET.get('usuario')
    tipo = request.GET.get('tipo')
    desde = request.GET.get('desde')
    hasta = request.GET.get('hasta')

    if usuario:
        queryset = queryset.filter(usuario__username=usuario)
    if tipo:
        queryset = queryset.filter(tipo=tipo)
    if desde:
        queryset = queryset.filter(fecha_hora__date__gte=parse_date(desde))
    if hasta:
        queryset = queryset.filter(fecha_hora__date__lte=parse_date(hasta))

    # Crear Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Actividades"
    ws.append(["Fecha", "Usuario", "Tipo", "Descripción"])

    for act in queryset:
        ws.append([
            act.fecha_hora.strftime("%Y-%m-%d %H:%M:%S"),
            act.usuario.username if act.usuario else "(anónimo)",
            act.get_tipo_display(),
            act.descripcion
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=actividades.xlsx'
    wb.save(response)
    return response



#------------------------------------------------------------------------------
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from django.http import HttpResponse
from .models import Actividad
from django.utils.dateparse import parse_date


@login_required
def exportar_actividades_pdf(request):
    queryset = Actividad.objects.select_related('usuario').order_by('-fecha_hora')

    usuario = request.GET.get('usuario')
    tipo = request.GET.get('tipo')
    desde = request.GET.get('desde')
    hasta = request.GET.get('hasta')

    if usuario:
        queryset = queryset.filter(usuario__username=usuario)
    if tipo:
        queryset = queryset.filter(tipo=tipo)
    if desde:
        queryset = queryset.filter(fecha_hora__date__gte=parse_date(desde))
    if hasta:
        queryset = queryset.filter(fecha_hora__date__lte=parse_date(hasta))

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=actividades.pdf'

    p = canvas.Canvas(response, pagesize=A4)
    width, height = A4
    y = height - 50

    p.setFont("Helvetica-Bold", 14)
    p.drawString(40, y, "Historial de Actividades")
    y -= 30

    p.setFont("Helvetica", 10)
    for actividad in queryset:
        texto = f"{actividad.fecha_hora.strftime('%Y-%m-%d %H:%M:%S')} | {actividad.usuario.username if actividad.usuario else '(anónimo)'} | {actividad.get_tipo_display()} | {actividad.descripcion}"
        p.drawString(40, y, texto)
        y -= 20
        if y < 50:
            p.showPage()
            p.setFont("Helvetica", 10)
            y = height - 50

    p.save()
    return response

#---------------------------------------------------------------------------------

@csrf_exempt
def registrar_clic_link(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            url = data.get('url', 'URL desconocida')
            log_actividad(request, TipoActividad.CLIC_LINK, f"Click en link: {url}")
            return JsonResponse({'status': 'ok'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'mensaje': str(e)}, status=400)
    return JsonResponse({'status': 'método no permitido'}, status=405)




#---------------------------------------------------------------------------------





# ViewSet para usuarios con perfil
class UserViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = User.objects.all().select_related('userprofile')
    serializer_class = UserSerializer

# Diario Digital
class DiarioDigitalViewSet(viewsets.ModelViewSet):
    queryset = DiarioDigital.objects.all()
    serializer_class = DiarioDigitalSerializer

# Categorias
class CategoriaViewSet(viewsets.ModelViewSet):
    queryset = Categoria.objects.all()
    serializer_class = CategoriaSerializer

# Links Relevantes


class LinkRelevanteViewSet(viewsets.ModelViewSet):
    queryset = LinkRelevante.objects.select_related('cargado_por', 'diario_digital').prefetch_related('categorias')
    permission_classes = [IsAuthenticated]

    def get_serializer_class(self):
        if self.action in ['create', 'update', 'partial_update']:
            return LinkRelevanteCreateSerializer  # usar create serializer para escritura
        return LinkRelevanteSerializer  # por defecto, para lectura

    def get_queryset(self):
        queryset = super().get_queryset()
        estado = self.request.query_params.get('estado')
        fecha_inicio = self.request.query_params.get('fecha_inicio')
        fecha_fin = self.request.query_params.get('fecha_fin')

        if estado:
            queryset = queryset.filter(estado=estado)
        if fecha_inicio:
            queryset = queryset.filter(fecha_carga__date__gte=parse_date(fecha_inicio))
        if fecha_fin:
            queryset = queryset.filter(fecha_carga__date__lte=parse_date(fecha_fin))

        return queryset



# Artículos
class ArticuloViewSet(viewsets.ModelViewSet):
    queryset = Articulo.objects.select_related('generado_por').prefetch_related('links_incluidos')
    serializer_class = ArticuloSerializer





class ActividadViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = Actividad.objects.select_related('usuario').order_by('-fecha_hora')
    serializer_class = ActividadSerializer

    def get_queryset(self):
        queryset = super().get_queryset()
        usuario = self.request.query_params.get('usuario')
        tipo = self.request.query_params.get('tipo')
        desde = self.request.query_params.get('desde')
        hasta = self.request.query_params.get('hasta')

        if usuario:
            queryset = queryset.filter(usuario__username=usuario)
        if tipo:
            queryset = queryset.filter(tipo=tipo)
        if desde:
            queryset = queryset.filter(fecha_hora__date__gte=parse_date(desde))
        if hasta:
            queryset = queryset.filter(fecha_hora__date__lte=parse_date(hasta))

        return queryset

